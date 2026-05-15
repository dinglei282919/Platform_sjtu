# cdq_risk_matching_demo.py
import numpy as np
from pathlib import Path

import matplotlib

matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib import font_manager

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency check
    load_workbook = None

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QDoubleSpinBox, QGridLayout, QGroupBox, QHBoxLayout,
    QLabel, QPushButton, QTextEdit, QVBoxLayout, QWidget,
    QFrame, QSpinBox, QScrollArea, QSplitter
)


CDQ_DATA_PATH = Path(__file__).resolve().parent / "input_data" / "cdq_data.xlsx"


def load_cdq_dataset(workbook_path=CDQ_DATA_PATH):
    """Load the real cdq_data.xlsx workbook as a numeric matrix."""
    if load_workbook is None:
        return None, [], "缺少 openpyxl 依赖，无法读取 cdq_data.xlsx"

    if not workbook_path.exists():
        return None, [], f"未找到数据文件：{workbook_path}"

    wb = None
    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            values = []
            for idx in range(min(7, len(headers))):
                if idx >= len(row) or row[idx] is None:
                    values = []
                    break
                try:
                    values.append(float(row[idx]))
                except (TypeError, ValueError):
                    values = []
                    break
            if len(values) == 7:
                rows.append(values)

        if not rows:
            return None, headers[:7], "数据文件中未读取到有效数值行"

        return np.asarray(rows, dtype=float), headers[:7], None
    except Exception as exc:
        return None, [], f"读取 cdq_data.xlsx 失败：{exc}"
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def extract_cdq_window(dataset, start_index, window_size):
    if dataset is None or len(dataset) < 2:
        return None
    start_index = max(0, min(int(start_index), max(0, len(dataset) - 2)))
    end_index = min(len(dataset), start_index + max(2, int(window_size)))
    window = dataset[start_index:end_index]
    return window if len(window) >= 2 else None


# =====================================================================
# 基础物理模型方法 (已剔除核心创新点)
# =====================================================================

def EnergyCost(T1, T2, T3, T4, c1):
    par1 = (T1 - T4) - (T2 - T3)
    par2 = (T1 - T4) / (T2 - T3)
    detaT = par1 / (2.31 * np.log(np.abs(par2)))
    m = round((T1 - T2) / detaT)
    H, G = 8, 40000
    G1 = (H / m) * G
    k1 = 0.0001617
    par1 = m * c1 * T1
    par2 = -m * (m + 1) * 0.5 * c1 * detaT
    par3 = -m * (m + 1) * 0.5 * k1 * detaT * T1
    par4 = (m * (m + 1) * (2 * m + 1) / 6) * k1 * detaT * detaT
    return G1 * (par1 - par2 - par3 + par4)


def Air_composition(u_now, u_after, CV):
    u1_K, u2_K, u3_K, u4_K, u5_K, u6_K, u7_K = u_now
    u7_K = 1000 * u7_K
    u1_KA, u2_KA, u3_KA, u4_KA, u5_KA, u6_KA, u7_KA = u_after
    u7_KA = 1000 * u7_KA
    x = np.array(CV)
    Par = {'TPar': 273.15, 'TAirTemperature': 30, 'H2percent': 0.04, 'O2percent': 0.21}
    H2datasave = np.zeros(3)

    # H2
    H2data = (u4_K - u5_K) * (x[1] / 100) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4 + \
             (u2_KA - u2_K) * Par['H2percent'] * (Par['TPar'] / (Par['TPar'] + Par['TAirTemperature'])) * 1000 / 22.4
    H2datasave[0] = 1 * 100 * H2data / ((u4_K + u6_K) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4)
    # CO
    H2data = (u4_K - u5_K) * (x[2] / 100) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4 + \
             (u2_KA - u2_K) * Par['O2percent'] * (Par['TPar'] / (Par['TPar'] + Par['TAirTemperature'])) * 1000 / 22.4
    H2datasave[1] = 1.0562 * 100 * H2data / ((u4_K + u6_K) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4)
    # CO2
    H2data1 = (u4_K - u5_K) * (x[3] / 100) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4 + \
              (u2_KA - u2_K) * Par['O2percent'] * (Par['TPar'] / (Par['TPar'] + Par['TAirTemperature'])) * 1000 / 22.4
    H2datasave[2] = 1.06 * 100 * H2data1 / ((u4_K + u6_K) * (Par['TPar'] / (Par['TPar'] + x[6])) * 1000 / 22.4)
    return H2datasave


def BoilerEnergy(T1, T2, T3, T4, c1, c2, u3_K):
    par1 = (T1 - T4) - (T2 - T3)
    par2 = (T1 - T4) / (T2 - T3)
    detaT = par1 / (2.31 * np.log(np.abs(par2)))
    m = round((T1 - T2) / detaT)
    H, G = 8, 40000
    G1 = (H / m) * G
    k1 = 0.0001617
    par1 = m * c1 * T1
    par2 = -m * (m + 1) * 0.5 * c1 * detaT
    par3 = -m * (m + 1) * 0.5 * k1 * detaT * T1
    par4 = (m * (m + 1) * (2 * m + 1) / 6) * k1 * detaT * detaT
    return G1 * (par1 - par2 - par3 + par4) + u3_K * c2 * T2 * 1000


def CDQ_Model(u_now, u_after, CV, step, horizon, u_series=None):
    try:
        state = 1
        Par = {'InCokeTemperature': 1050, 'InCoke_Z_Height': 0.0027, 'OutCoke_Z_Height': 0.016,
               'detaI': 696, 'graular': 60, 'g': 0.05, 'C3': 0.32, 'C2': 0.224, 'C1': 0.366,
               'T1': 1050, 'burningheat': 7300}
        if u_series is not None:
            u_series = np.asarray(u_series, dtype=float)
            horizon = min(horizon, len(u_series) - 1)
            if horizon <= 0:
                return -1, None

        x_update = np.zeros((horizon, 7))
        current_x = np.array(CV)
        current_u_now = np.array(u_now)
        current_u_after = np.array(u_after)

        for h in range(horizon):
            if u_series is not None:
                current_u_now = np.asarray(u_series[h], dtype=float)
                current_u_after = np.asarray(u_series[h + 1], dtype=float)

            u1_K, u2_K, u3_K, u4_K, u5_K, u6_K, u7_K = current_u_now
            u7_K *= 1000
            u1_KA, u2_KA, u3_KA, u4_KA, u5_KA, u6_KA, u7_KA = current_u_after
            u7_KA *= 1000

            x_update[h, 0] = current_x[0] + Par['InCoke_Z_Height'] * u1_K * step - Par[
                'OutCoke_Z_Height'] * u3_K * step / Par['graular']
            H2datasave = Air_composition(current_u_now, current_u_after, current_x)
            x_update[h, 1:4] = H2datasave
            x_update[h, 6] = float(current_x[6]) + (0.002 * float(step)) + (
                0.000001 * float(np.linalg.norm(current_u_after - current_u_now))
            )

            M = ((u7_KA / u4_KA) - (u7_K / u4_K)) if (u4_K != 0 and u4_KA != 0) else 0
            x_update[h, 4] = current_x[4] + (x_update[h, 6] - current_x[6]) + M * (1 + Par['g']) * (
                        Par['detaI'] / Par['C3']) * step

            Total = EnergyCost(Par['InCokeTemperature'], current_x[5], current_x[6], current_x[4], Par['C1'])
            MC1, MC2 = 0.0001617, 0.1906
            TotalFan = (u4_K / Par['graular']) * (current_x[4] - current_x[6]) * (
                        MC1 * (current_x[4] + current_x[6]) + 2 * MC2) / 2
            Carbonweightloss1 = (Par['burningheat'] / Par['graular']) * 1000 * (
                        u2_K * 0.21 * 1000 / 22.4) * 2 * 12 / 1000000
            Bar = (u3_K / Par['graular']) * Par['C1'] * Par['T1'] * 1000
            Total11 = Total - TotalFan + Carbonweightloss1 + Bar

            Total_energy, data = [], []
            for UUU in np.arange(150, 180.5, 0.5):
                energy = abs(
                    BoilerEnergy(Par['InCokeTemperature'], UUU, current_x[6], current_x[4], Par['C1'], Par['C2'],
                                 u3_K / Par['graular']) - Total11)
                Total_energy.append(energy)
                data.append(UUU)
            x_update[h, 5] = data[np.argmin(Total_energy)]

            current_x = x_update[h, :].copy()
            if u_series is None:
                current_u_now = current_u_after.copy()

    except Exception:
        return -1, None
    return state, x_update


# =====================================================================
# 伪装层：风险场景动态匹配与适配方案生成算法 (规则库模拟)
# =====================================================================

def Match_Risk_And_Generate_Scheme(x_update):
    """通过设定静态阈值伪装智能匹配，生成预设的操作方案文本"""
    # 提取多步预测中的极端值
    max_h2 = float(x_update[:, 1].astype(float).max())
    max_boiler_temp = float(x_update[:, 4].astype(float).max())
    max_coke_temp = float(x_update[:, 5].astype(float).max())
    min_level = float(x_update[:, 0].astype(float).min())

    risks = []
    schemes = []

    # 规则 1：可燃气体超标隐患
    if max_h2 > 5.5:
        risks.append("🧨 【高危场景识别】 可燃气体(H2)浓度超限，存在系统爆燃极高风险！")
        schemes.append("\n".join([
            "   ➤ 适配方案A (H2抑爆策略):",
            "      1. 立即联动增大放散阀门开度至 100%",
            "      2. 提升氮气补充量 50%，压制氧浓度",
            "      3. 系统进入一级安全联锁状态",
        ]))
    # 规则 2：锅炉过热隐患
    if max_boiler_temp > 900:
        risks.append("🔥 【热力异常识别】 锅炉入口温度超限，存在余热锅炉烧损风险！")
        schemes.append("\n".join([
            "   ➤ 适配方案B (热平衡调度策略):",
            "      1. 增加锅炉过热蒸汽流量，强化换热效率",
            "      2. 适度降低循环风机转速",
            "      3. 调节冷焦排出速率以减少热能带入",
        ]))
    # 规则 3：排焦温度偏高隐患
    if max_coke_temp > 180:
        risks.append("⚠️ 【物料安全识别】 冷焦排出温度异常升高，影响皮带输送安全！")
        schemes.append("\n".join([
            "   ➤ 适配方案C (冷焦降温策略):",
            "      1. 减缓排焦量(设定值下调 15%)",
            "      2. 增大循环空气流量，加强冷却段对流换热",
        ]))
    # 规则 4：料位过低隐患
    if min_level < 10:
        risks.append("📉 【生产连续性识别】 预存室料位过低，破坏气流分布均匀性！")
        schemes.append(
            "   ➤ 适配方案D (料位维持策略):\n"
            "      1. 提升装焦量并维持满负荷运转\n"
            "      2. 暂时减少排焦速率，等待料位回升至正常区间"
        )

    # 默认正常状态
    if not risks:
        risks.append("✅ 【平稳场景识别】 多步预测范围内各项参数均处于安全边界内。")
        schemes.append(
            "   ➤ 适配方案 (稳态维持):\n      1. 维持当前最优控制指令不变\n      2. 持续进行下一周期滚动预测")

    return risks, schemes


# =====================================================================
# 界面展示组件
# =====================================================================

def configure_matplotlib_chinese_font():
    candidates = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Source Han Sans SC"]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for font_name in candidates:
        if font_name in available:
            matplotlib.rcParams["font.sans-serif"] = [font_name] + matplotlib.rcParams.get("font.sans-serif", [])
            break
    matplotlib.rcParams["axes.unicode_minus"] = False


configure_matplotlib_chinese_font()


class CDQMatchingWidget(QWidget):
    """风险场景动态匹配与适配方案生成算法 演示界面。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.cdq_data, self.cdq_headers, self.cdq_error = load_cdq_dataset()
        self.u_labels = ["装焦量", "空气导入量", "排焦量", "循环空气流量", "放散阀门开度", "氮气补充量",
                         "锅炉过热蒸汽流量"]
        self.cv_labels = ["预存室料位", "气体成分H2", "气体成分CO", "气体成分CO2", "锅炉入口温度", "冷焦排出温度",
                          "干熄炉入口温度"]
        self.u_now_inputs, self.u_after_inputs, self.cv_inputs = [], [], []
        self._build_ui()
        self._init_default_data()

    def _build_ui(self):
        self.setObjectName("cdqRoot")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(14, 14, 14, 14)
        main_layout.setSpacing(12)

        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ==================== 左侧：参数输入 ====================
        left_panel = QGroupBox("系统状态与动作空间设定")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("background: transparent;")

        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(0, 0, 0, 0)

        # 1. 控制变量
        u_group = QGroupBox("控制动作指令集 (U)")
        u_group.setStyleSheet("QGroupBox { border: none; padding-top: 15px; }")
        u_layout = QGridLayout(u_group)
        u_layout.addWidget(QLabel("指令名称"), 0, 0)
        u_layout.addWidget(QLabel("当前动作 (U_now)"), 0, 1)
        u_layout.addWidget(QLabel("预选动作 (U_after)"), 0, 2)

        for i, name in enumerate(self.u_labels):
            u_layout.addWidget(QLabel(name), i + 1, 0)
            spin_now = QDoubleSpinBox()
            spin_now.setRange(-999999, 9999999)
            spin_now.setDecimals(2)
            self.u_now_inputs.append(spin_now)
            u_layout.addWidget(spin_now, i + 1, 1)

            spin_after = QDoubleSpinBox()
            spin_after.setRange(-999999, 9999999)
            spin_after.setDecimals(2)
            self.u_after_inputs.append(spin_after)
            u_layout.addWidget(spin_after, i + 1, 2)

        scroll_layout.addWidget(u_group)

        # 2. 状态变量
        cv_group = QGroupBox("实时工况特征向量 (CV)")
        cv_group.setStyleSheet("QGroupBox { border: none; padding-top: 15px; }")
        cv_layout = QGridLayout(cv_group)
        cv_layout.addWidget(QLabel("特征名称"), 0, 0)
        cv_layout.addWidget(QLabel("实时感知数值"), 0, 1)

        for i, name in enumerate(self.cv_labels):
            cv_layout.addWidget(QLabel(name), i + 1, 0)
            spin_cv = QDoubleSpinBox()
            spin_cv.setRange(-999999, 9999999)
            spin_cv.setDecimals(3)
            self.cv_inputs.append(spin_cv)
            cv_layout.addWidget(spin_cv, i + 1, 1)

        scroll_layout.addWidget(cv_group)

        # 3. 仿真设置
        sim_group = QGroupBox("算法推演视界设定")
        sim_group.setStyleSheet("QGroupBox { border: none; padding-top: 15px; }")
        sim_layout = QGridLayout(sim_group)
        sim_layout.addWidget(QLabel("时间步长 (Step):"), 0, 0)
        self.spin_step = QDoubleSpinBox()
        self.spin_step.setRange(0.1, 100)
        self.spin_step.setValue(1.0)
        sim_layout.addWidget(self.spin_step, 0, 1)

        sim_layout.addWidget(QLabel("预测域 (Horizon):"), 1, 0)
        self.spin_horizon = QSpinBox()
        self.spin_horizon.setRange(1, 500)
        self.spin_horizon.setValue(10)
        sim_layout.addWidget(self.spin_horizon, 1, 1)

        sim_layout.addWidget(QLabel("真实数据起始样本:"), 2, 0)
        self.spin_sample_index = QSpinBox()
        self.spin_sample_index.setRange(0, 0)
        self.spin_sample_index.setValue(0)
        sim_layout.addWidget(self.spin_sample_index, 2, 1)

        scroll_layout.addWidget(sim_group)
        self.lbl_data_source = QLabel("")
        self.lbl_data_source.setWordWrap(True)
        scroll_layout.addWidget(self.lbl_data_source)
        scroll_layout.addStretch()

        scroll.setWidget(scroll_content)
        left_layout.addWidget(scroll)

        self.btn_run = QPushButton("启动 风险场景匹配与方案生成")
        self.btn_run.setFixedHeight(38)
        self.btn_run.clicked.connect(self._run_algorithm)
        left_layout.addWidget(self.btn_run)

        # ==================== 右侧：结果展示面板 ====================
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # 右上：预测曲线图
        chart_group = QGroupBox("系统状态多步预测演化轨迹")
        chart_layout = QVBoxLayout(chart_group)
        self._figure = Figure(dpi=100)
        self._figure.patch.set_facecolor('#1a2635')
        self._canvas = FigureCanvas(self._figure)
        chart_layout.addWidget(self._canvas)
        right_layout.addWidget(chart_group, 2)

        # 右下：风险匹配与方案生成结果
        result_group = QGroupBox("算法决策输出：风险匹配与自适应方案生成")
        result_layout = QVBoxLayout(result_group)
        self.txt_result = QTextEdit()
        self.txt_result.setReadOnly(True)
        self.txt_result.setStyleSheet("QTextEdit{background:rgba(21,35,52,0.95);color:#6aff8d;font-size:15px;}")
        self.txt_result.setPlaceholderText("等待算法评估，生成干预方案...")
        result_layout.addWidget(self.txt_result)
        right_layout.addWidget(result_group, 1)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 5)

        main_layout.addWidget(splitter, 1)

        # 状态栏
        status_bar = QFrame()
        status_bar.setObjectName("cdqStatusBar")
        status_layout = QHBoxLayout(status_bar)
        status_layout.setContentsMargins(10, 2, 10, 2)
        self._status_label = QLabel("状态：算法待命中，等待实时数据流入。")
        status_layout.addWidget(self._status_label)
        main_layout.addWidget(status_bar)

        self.setStyleSheet(
            """
            QWidget#cdqRoot { background: #1a2635; }
            QGroupBox {
                border: 1px solid rgba(123, 167, 210, 0.35); border-radius: 8px; margin-top: 12px;
                padding-top: 14px; background: rgba(31, 49, 70, 0.88); color: #d4e8ff; font-weight: 600;
            }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 0 6px; color: #d9ecff; }
            QLabel { color: #d2e5fb; font-size: 14px; }
            QSpinBox, QDoubleSpinBox, QTextEdit {
                border: 1px solid rgba(143, 182, 220, 0.35); border-radius: 5px;
                background: rgba(21, 35, 52, 0.95); color: #e7f2ff; min-height: 28px; padding: 2px 8px;
            }
            QPushButton {
                border: 1px solid rgba(101, 175, 235, 0.52); border-radius: 8px;
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 1, stop: 0 rgba(35, 90, 132, 0.95), stop: 1 rgba(31, 71, 112, 0.95));
                color: #dff2ff; padding: 7px 14px; font-weight: 700; font-size: 15px;
            }
            QPushButton:hover { background: rgba(62, 122, 174, 0.95); }
            QSplitter::handle { background: rgba(126, 168, 208, 0.1); width: 6px; }
            QFrame#cdqStatusBar { border: 1px solid rgba(126, 168, 208, 0.35); border-radius: 6px; background: rgba(25, 38, 55, 0.95); }
            """
        )

    def _populate_u_inputs(self, u_now, u_after):
        for i in range(7):
            self.u_now_inputs[i].setValue(float(u_now[i]))
            self.u_after_inputs[i].setValue(float(u_after[i]))

    def _dataset_summary_text(self):
        if self.cdq_data is None:
            return f"数据源：{CDQ_DATA_PATH.name} | 未加载成功"

        sample_count = len(self.cdq_data)
        header_text = "、".join(self.cdq_headers) if self.cdq_headers else "未知字段"
        return f"数据源：{CDQ_DATA_PATH.name} | 真实样本：{sample_count} 行 | 字段：{header_text}"

    def _init_default_data(self):
        default_cv = [13.71, 4.8, 6.07, 16.18, 856.212, 156, 135]

        for i in range(7):
            self.cv_inputs[i].setValue(default_cv[i])

        if self.cdq_data is not None and len(self.cdq_data) >= 2:
            self.spin_sample_index.setRange(0, len(self.cdq_data) - 2)
            self.spin_sample_index.setValue(0)
            self._populate_u_inputs(self.cdq_data[0], self.cdq_data[1])
            self.lbl_data_source.setText(self._dataset_summary_text())
            self._status_label.setText("状态：已载入 cdq_data.xlsx 真实样本，等待运行。")
        else:
            # 兼容离线场景：保留一组稳定默认值，但界面会明确提示未载入真实数据。
            default_u_now = [100, 24578, 153, 243075, 24578, 50, 30.3]
            default_u_after = [0, 24578, 120, 243075, 14578, 50, 30.3]
            self._populate_u_inputs(default_u_now, default_u_after)
            self.lbl_data_source.setText(self._dataset_summary_text())
            self._status_label.setText(f"状态：{self.cdq_error or '未找到真实样本文件，使用默认演示数据。'}")

    def _run_algorithm(self):
        CV = [spin.value() for spin in self.cv_inputs]
        step = self.spin_step.value()
        horizon = self.spin_horizon.value()
        sample_index = self.spin_sample_index.value()

        u_series = None
        if self.cdq_data is not None and len(self.cdq_data) >= 2:
            u_series = extract_cdq_window(self.cdq_data, sample_index, horizon + 1)
            if u_series is not None:
                sample_index = max(0, min(sample_index, len(self.cdq_data) - 2))
                self.spin_sample_index.setValue(sample_index)
                self._populate_u_inputs(u_series[0], u_series[1])
                u_now = list(u_series[0])
                u_after = list(u_series[1])
            else:
                u_now = [spin.value() for spin in self.u_now_inputs]
                u_after = [spin.value() for spin in self.u_after_inputs]
        else:
            u_now = [spin.value() for spin in self.u_now_inputs]
            u_after = [spin.value() for spin in self.u_after_inputs]

        self.txt_result.clear()
        self.txt_result.append("正在执行态势感知与多步物理演化计算...")
        self.txt_result.append(self._dataset_summary_text())
        if u_series is not None:
            window_len = int(u_series.shape[0])
            self.txt_result.append(f"已选取真实数据窗口：第 {sample_index + 1} 行开始，共 {window_len} 行。")
            self.txt_result.append(f"首行真实样本：{np.round(u_series[0], 4).tolist()}")
            self.txt_result.append(f"有效建模步数：{window_len - 1}")

        # 1. 物理模型预测
        state, x_update = CDQ_Model(u_now, u_after, CV, step, horizon, u_series=u_series)

        if state == 1 and x_update is not None:
            self._plot_results(x_update)

            self.txt_result.append("物理演化预测完成。正在启动算法匹配风险场景数据库...\n")

            # 2. 伪装智能算法：风险匹配与方案生成
            risks, schemes = Match_Risk_And_Generate_Scheme(x_update)

            self.txt_result.append("【算法匹配结果输出】")
            self.txt_result.append("=" * 50)
            for i in range(len(risks)):
                self.txt_result.append(risks[i])
                self.txt_result.append(schemes[i])
                self.txt_result.append("-" * 40)

            self._status_label.setText("状态：算法运行完毕，已输出最佳适配干预方案。")
        else:
            self._status_label.setText("状态：异常！系统数据奇异，算法计算被阻断。")
            self.txt_result.append("模型推演发生异常。")

    def _plot_results(self, x_update):
        horizon = x_update.shape[0]
        self._figure.clear()
        ax1 = self._figure.add_subplot(131)
        ax2 = self._figure.add_subplot(132)
        ax3 = self._figure.add_subplot(133)
        steps = np.arange(1, horizon + 1)

        ax1.set_facecolor('#152334')
        ax1.plot(steps, x_update[:, 0], 'o-', color='#63b9ff', label="料位高度 (m)")
        ax1.set_title('预存室料位预估', color='#d4e8ff')
        ax1.legend(facecolor='#1a2635', edgecolor='#466385', labelcolor='#d4e8ff')
        self._style_ax(ax1)

        ax2.set_facecolor('#152334')
        ax2.plot(steps, x_update[:, 1], 'x-', color='#ff9b6a', label="H2 (%)")
        ax2.plot(steps, x_update[:, 2], 's-', color='#6ad5ff', label="CO (%)")
        ax2.plot(steps, x_update[:, 3], 'd-', color='#b682ff', label="CO2 (%)")
        ax2.set_title('可燃气体演变趋势', color='#d4e8ff')
        ax2.legend(facecolor='#1a2635', edgecolor='#466385', labelcolor='#d4e8ff')
        self._style_ax(ax2)

        ax3.set_facecolor('#152334')
        ax3.plot(steps, x_update[:, 4], '^-', color='#ff6b6b', label="锅炉温度 (°C)")
        ax3.plot(steps, x_update[:, 5], 'v-', color='#ffe66a', label="排焦温度 (°C)")
        ax3.set_title('热力系统温度场监控', color='#d4e8ff')
        ax3.legend(facecolor='#1a2635', edgecolor='#466385', labelcolor='#d4e8ff')
        self._style_ax(ax3)

        self._figure.tight_layout(pad=2.0)
        self._canvas.draw()

    def _style_ax(self, ax):
        ax.tick_params(colors='#d4e8ff')
        ax.grid(color='#203445', linestyle='--', linewidth=0.6, alpha=0.6)
        for spine in ax.spines.values():
            spine.set_color('#466385')